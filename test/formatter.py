from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font,  Alignment

wb = load_workbook('Grades.xlsx')
ws = wb.active

#bold + center the headers
for cell in ws[1]:
    cell.font = Font(bold = True)
    cell.alignment = Alignment(horizontal='center')

#color codes by grades
colors = {"A": "C6EFCE", "B": "FFEB9C", "C": "FFC7CE"}  # green, yellow, red

for row in ws.iter_rows(min_row = 2,max_row = ws.max_row):
    grade = row[-1].value 
    fill = PatternFill(start_color = colors[grade],end_color = colors[grade], fill_type='solid')
    for cell in row:
        cell.fill = fill

wb.save('Grades_formatted.xlsx')
print ('formatted file saved!!')
    