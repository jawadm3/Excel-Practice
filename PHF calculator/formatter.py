from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = load_workbook('phf_raw.xlsx')
ws = wb.active

col_widths = [14,10,10,10,10,14,13,8,10]
for i,width in enumerate(col_widths,start=1):
    ws.column_dimensions[ws.cell(1,i).column_letter].width = width

# --- Header row: dark blue background, white bold text, centered --
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type="solid") 
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill   = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# --- Data rows ---
ok_fill = PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid') #green
peaky_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') #red
thin_side = Side(style='thin')
border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
    status = row[-1].value
    fill = ok_fill if status == "ok" else peaky_fill

    for cell in row:
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row[0].alignment = Alignment(horizontal='left') #approach column left-aligned

ws.freeze_panes = 'A2' #freezing the header row

wb.save('phf_report.xlsx')
print('report saved')