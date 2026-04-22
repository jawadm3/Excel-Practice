from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

print('\nDo you want to create a new workbook or edit an existing one?')
option = input('Type "new" or "existing": ').strip().lower()


def newinputs(rows, columns):
    wb = Workbook()
    sheetname = input('Enter the name of the sheet: ')
    wb.active.title = sheetname
    sheet1 = wb.active

    print('\nEnter column headings:')
    for j in range(1, columns + 1):
        heading = input(f'Heading for column {j}: ')
        sheet1.cell(row=1, column=j).value = heading

    print('\nEnter values for each cell:')
    for i in range(2, rows + 2):
        for j in range(1, columns + 1):
            value = input(f'Value at ({i-1}, {j}): ')
            sheet1.cell(row=i, column=j).value = value

    filename = input('\nEnter filename to save (with .xlsx): ')
    wb.save(filename)
    print(f'\nWorkbook saved as {filename}')


def edit_existing(path):
    wb = load_workbook(path)
    sheetname = input(f'\nSheets available: {wb.sheetnames}\nWhich sheet do you want to edit? ')
    sheet1 = wb[sheetname]

    rows = sheet1.max_row
    columns = sheet1.max_column

    print(f'\nThis sheet has {rows} rows and {columns} columns.\n')
    print('Current data:\n')

    for i in range(1, rows + 1):
        row_values = [sheet1.cell(i, j).value for j in range(1, columns + 1)]
        print(row_values)

    print('\nHow many new rows do you want to add?')
    new_rows = int(input('Enter number: '))

    start_row = rows + 1

    for i in range(start_row, start_row + new_rows):
        print(f'\nEntering values for new row {i - 1}:')
        for j in range(1, columns + 1):
            value = input(f'Value at ({i}, {j}): ')
            sheet1.cell(row=i, column=j).value = value

    wb.save(path)
    print(f'\nWorkbook updated and saved: {path}')


# MAIN LOGIC
if option == 'new':
    rows = int(input('How many data rows do you want? '))
    columns = int(input('How many columns do you want? '))
    newinputs(rows, columns)

elif option == 'existing':
    workbook_name = input('Enter the workbook path + name: ')
    edit_existing(workbook_name)

else:
    print('Invalid option. Type "new" or "existing".')
