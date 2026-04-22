import openpyxl as xl
workbook = xl.load_workbook('raw_data.xlsx')
sheets = workbook.sheetnames

print(f'\nList of the sheets in the workbook ', sheets, end='\n\n')
print(f'Active sheet in the workbook ',workbook.active.title, end='\n\n')

print(f'Value at A1 {workbook['Sheet1']['A1'].value}', end='\n\n')

print (f'value at 3rd row and 1st col',workbook['Sheet1'].cell(3,1).value, end='\n\n')

sheet1 = workbook['Sheet1']
rows = sheet1.max_row
columns = sheet1.max_column
print(f'number of rows ',rows,'number of columns ',  columns, end='\n\n')



for i in range (1,rows+1):
    for j in range (1,columns+1):
        print(f'value at {i, j}: {sheet1.cell(i,j).value}', end='\n\n')

